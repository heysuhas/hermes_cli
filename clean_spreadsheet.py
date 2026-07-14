import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_and_format_spreadsheet(input_path, output_path):
    print(f"Reading input file: {input_path}")
    # 1. Load Data
    df = pd.read_csv(input_path)
    
    # 2. Clean Data
    # Strip whitespace from string columns
    for col in df.select_dtypes(include='object').columns:
        df[col] = df[col].astype(str).str.strip()
        
    initial_rows = len(df)
    
    # Capitalize names and products
    if 'customer_name' in df.columns:
        df['customer_name'] = df['customer_name'].str.title()
    if 'Product' in df.columns:
        df['Product'] = df['Product'].str.capitalize()
        
    # Standardize dates
    if 'date' in df.columns:
        df['date'] = df['date'].astype(str).str.replace('.', '-', regex=False).str.strip()
        parsed_dates = pd.to_datetime(df['date'], errors='coerce', format='mixed')
        df['date'] = parsed_dates.dt.strftime('%Y-%m-%d').fillna('N/A')
        
    # Standardize currency and amount columns
    if 'amount' in df.columns:
        df['amount'] = df['amount'].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False).str.strip()
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
        
    # Standardize status
    if 'status' in df.columns:
        df['status'] = df['status'].fillna('Pending').astype(str).str.strip().str.capitalize()
        df['status'] = df['status'].replace({'': 'Pending', 'Nan': 'Pending', 'N/a': 'Pending'})
        
    # Deduplicate rows
    df = df.drop_duplicates()
    removed_dupes = initial_rows - len(df)
    print(f"Removed {removed_dupes} duplicate rows.")
    
    # Clean notes
    if 'notes' in df.columns:
        df['notes'] = df['notes'].fillna('').replace({'nan': '', 'N/a': ''})
        
    # Save clean dataframe to Excel
    df.to_excel(output_path, index=False)
    print(f"Saved cleaned data to Excel at: {output_path}")
    
    # 3. Apply professional styles using openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Ensure gridlines are visible and freeze top row
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = 'A2'
    
    # Define styles
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    white_bold_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=11, bold=False, color="000000")
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Format Header Row
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Format Data Rows
    for row_num in range(2, ws.max_row + 1):
        is_even = (row_num % 2 == 0)
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            
            # Zebra striping
            if is_even:
                cell.fill = zebra_fill
                
            # Alignment & number formatting based on column name
            col_name = df.columns[col_num - 1]
            if col_name == 'date':
                cell.alignment = Alignment(horizontal="center")
            elif col_name == 'status':
                cell.alignment = Alignment(horizontal="center")
            elif col_name == 'amount':
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
    # Add a Summary / Total Row
    total_row_idx = ws.max_row + 2  # Leave one blank row before Total
    ws.cell(row=total_row_idx, column=1, value="Total").font = Font(name="Segoe UI", size=11, bold=True)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="left")
    
    # Insert sum formula for Amount
    amount_col_idx = df.columns.get_loc('amount') + 1
    total_cell = ws.cell(row=total_row_idx, column=amount_col_idx, value=f"=SUM(D2:D{total_row_idx-2})")
    total_cell.font = Font(name="Segoe UI", size=11, bold=True)
    total_cell.number_format = '"$"#,##0.00'
    total_cell.alignment = Alignment(horizontal="right")
    
    # Double bottom border (accounting style)
    double_bottom_border = Border(
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='double', color='1B365D')
    )
    for col_num in range(1, ws.max_column + 1):
        ws.cell(row=total_row_idx, column=col_num).border = double_bottom_border
        
    # Auto-adjust column widths with extra padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Handle formulas or numbers length representation
                val_str = str(cell.value)
                if val_str.startswith('='):
                    val_str = "$123,456.78"  # estimate for Total representation
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output_path)
    print(f"Spreadsheet styled and saved to: {output_path}")

if __name__ == "__main__":
    clean_and_format_spreadsheet("dirty_sales_data.csv", "cleaned_sales_data.xlsx")
