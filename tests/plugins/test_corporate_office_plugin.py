from __future__ import annotations

import json

import pytest

from agent import corporate_policy as policy_module
from agent.corporate_policy import CorporatePolicy
from agent.corporate_path_access import clear_session_path_grants
from plugins.corporate_office import tools
from tools.terminal_tool import set_approval_callback


@pytest.fixture(autouse=True)
def reset_path_approval():
    clear_session_path_grants()
    set_approval_callback(None)
    yield
    clear_session_path_grants()
    set_approval_callback(None)


def test_excel_plan_apply_backup_and_verify(monkeypatch, tmp_path):
    from openpyxl import Workbook, load_workbook

    source = tmp_path / "report.xlsx"
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook["Data"]["A1"] = "before"
    workbook.save(source)
    workbook.close()

    monkeypatch.setattr(
        policy_module,
        "_policy_cache",
        CorporatePolicy(
            enabled=True,
            allowed_roots=(tmp_path,),
            audit_enabled=False,
        ),
    )
    monkeypatch.setattr(tools, "PLAN_DIR", tmp_path / "plans")
    monkeypatch.setattr(tools, "_ask_approval", lambda record: True)
    monkeypatch.setattr(
        tools,
        "_apply_with_com",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("force fallback")),
    )

    plan = json.loads(
        tools.handle_plan(
            {
                "path": str(source),
                "operations": [
                    {
                        "action": "set_formula",
                        "sheet": "Data",
                        "cell": "B2",
                        "formula": "=1+1",
                    }
                ],
            }
        )
    )
    assert plan["requires_approval"] is True

    applied = json.loads(tools.handle_apply({"plan_id": plan["plan_id"]}))
    assert applied["applied"] is True
    assert applied["engine"] == "openpyxl"
    assert applied["backup_path"]

    workbook = load_workbook(source, data_only=False)
    try:
        assert workbook["Data"]["B2"].value == "=1+1"
    finally:
        workbook.close()


def test_document_path_outside_approved_roots_is_blocked(monkeypatch, tmp_path):
    allowed = tmp_path / "allowed"
    allowed.mkdir()
    outside = tmp_path / "outside.docx"
    outside.write_bytes(b"not-a-real-docx")
    monkeypatch.setattr(
        policy_module,
        "_policy_cache",
        CorporatePolicy(enabled=True, allowed_roots=(allowed,), audit_enabled=False),
    )

    result = json.loads(tools.handle_inspect({"path": str(outside)}))

    assert "outside" in result["error"].lower()


def test_document_extract_can_prompt_and_continue(monkeypatch, tmp_path):
    from docx import Document

    allowed = tmp_path / "allowed"
    downloads = tmp_path / "Downloads"
    allowed.mkdir()
    downloads.mkdir()
    document_path = downloads / "report.docx"
    document = Document()
    document.add_paragraph("Quarterly Prism conference summary")
    document.save(document_path)
    monkeypatch.setattr(
        policy_module,
        "_policy_cache",
        CorporatePolicy(
            enabled=True,
            allowed_roots=(allowed,),
            allowed_root_parents=(tmp_path,),
            audit_enabled=False,
        ),
    )
    set_approval_callback(lambda *args, **kwargs: "once")

    result = json.loads(tools.handle_extract({"path": str(document_path)}))

    assert result["chunks"][0]["text"] == "Quarterly Prism conference summary"
